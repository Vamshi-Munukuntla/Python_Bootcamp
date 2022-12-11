from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, IconSetRule, DataBarRule

spreadsheet = load_workbook('sample_data.xlsx')
sheet = spreadsheet.active

# TODO - Conditional Formatting of Pattern Fill
# red_background = PatternFill(bgColor='00FF0000')
# differential_style = DifferentialStyle(fill=red_background)
# rule = Rule(type='expression', dxf=differential_style)
# rule.formula = ["$B1<70"]
# sheet.conditional_formatting.add("A1:B100", rule)


# TODO - ColorScale
# color_scale_rule = ColorScaleRule(start_type='min', start_color='00FF0000',
#                                   end_type='max', end_color='00FFFF99')

# color_scale_rule = ColorScaleRule(start_type='num',
#                                   start_value=60,
#                                   start_color='00FF0000',
#
#                                   mid_type='num',
#                                   mid_value=75,
#                                   mid_color='00FFFF99',
#
#                                   end_type='num',
#                                   end_value=100,
#                                   end_color='00969696')
#
# sheet.conditional_formatting.add('B2:B100', color_scale_rule)

# TODO - IconSet
# icon_set_rule = IconSetRule('3Arrows',
#                             'num',
#                             [60, 85, 99])
#
# sheet.conditional_formatting.add("B2:B100", icon_set_rule)

# TODO - DataBar
data_bar_rule = DataBarRule(start_type='num',
                            start_value=60,
                            end_type='num',
                            end_value=100,
                            color='0000FF00')

sheet.conditional_formatting.add("B2:B100", data_bar_rule)
spreadsheet.save('sample_data5.xlsx')
