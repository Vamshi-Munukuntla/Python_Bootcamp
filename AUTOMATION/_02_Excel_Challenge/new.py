# TODO 1 - Open and read the cells of an Excel document with openpyxl module
from openpyxl import load_workbook
import pprint

print('Opening Excel spreadsheet...')
workbook = load_workbook('transactions.xlsx')
sheet = workbook.active
supplier_data = {}
print("Reading rows from the Excel spreadsheet...")
for row_num in range(2, sheet.max_row+1):
    transaction_type = sheet['B'+str(row_num)].value
    supplier_name = sheet['C'+str(row_num)].value
    amount = sheet['D' + str(row_num)].value

    # TODO 2 - Calculate all the transaction amounts and store it in dictionary data structure
    # Make sure the key for transaction type exist
    supplier_data.setdefault(transaction_type, {})
    # Make sure the key for supplier exist in the dictionary
    supplier_data[transaction_type].setdefault(supplier_name, {"transaction_count": 0,
                                                               "amount": 0})

    # Increase the transaction count for each row by one
    supplier_data[transaction_type][supplier_name]['transaction_count'] += 1
    # Sum transaction amount
    supplier_data[transaction_type][supplier_name]['amount'] += int(amount)


# TODO 3 - Write the data structure to .py file using pprint module
print('Writing to the output file...')
with open('output.py', mode='w') as output_file:
    output_file.write('all_transactions = '+pprint.pformat(supplier_data))

print('The task is completed...')
