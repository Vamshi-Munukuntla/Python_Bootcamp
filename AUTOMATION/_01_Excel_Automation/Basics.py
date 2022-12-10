from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pprint

workbook = load_workbook('AmazonReviews.xlsx')
sheet = workbook.active
# cell = sheet['B2']
# print(cell.value)


# print(workbook.sheetnames)
# print(sheet.title)

# cell1 = sheet.cell(row=2, column=2)
# print(cell1.value)


# cell_range = sheet['A2':'C3']
# print(cell_range)

# column_range = sheet["A":"C"]
# print(column_range)

# row_range = sheet[1:3]
# print(row_range)

# for i in range(1, 6, 2):
#     print(sheet.cell(row=i, column=4).value)

# for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
#     print(row)

# for column in sheet.iter_cols(min_row=1, max_row=10, min_col=1, max_col=10, values_only=True):
#     print(column)

# for row in sheet.rows:
#     print(row)

# for columns in sheet.columns:
#     for column in columns:
#         print(column.value)

# for col in list(sheet.columns)[1]:
#     print(col.value)

# for cells in sheet['A1':'C10']:
#     for cell in cells:
#         print(cell.coordinate, cell.value)

# print(get_column_letter(sheet.max_column))
# print(column_index_from_string('J'))

for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    print(value)

reviews={}
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=8, values_only=True):
    review_id = row[0]
    review = {
        'profile_name': row[1],
        'title': row[4],
        'rating': row[6]
    }
    reviews[review_id] = review
# print(reviews)
pprint.pprint(reviews)
